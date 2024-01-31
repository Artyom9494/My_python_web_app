import sqlite3
from http.server import BaseHTTPRequestHandler, HTTPServer
from start_sql.script_sql import start_script, start_value_sql
from jinja2 import Environment, FileSystemLoader, select_autoescape
from urllib.parse import unquote_plus
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
import fitz

# Создание базы данных SQLite
with sqlite3.connect('web_database.db') as wbd:  # создание и заполненение базы
    cursor = wbd.cursor()
    cursor.executescript(start_script)  # выполняем все скрипты sql

# Создание таблицы users (если она не существует)
try:
    db = sqlite3.connect('web_database.db')
    cursor = db.cursor()
    id_table_users = 1
    cursor.execute("SELECT id FROM users WHERE id = ?", [id_table_users])
    if cursor.fetchone() is None:
        start_users_values = ['Иванов', 'Иван', 'Петрович', 0, 0, '+7 999 999 99 99', 'Iva@iva']

        cursor.execute(
            "INSERT INTO users(second_name,first_name,patronymic,region_id,city_id,phone,email) VALUES(?,?,?,?,?,?,?)",
            start_users_values)

        cursor.executescript(start_value_sql)

        db.commit()
    else:
        print('такой id уже есть')
except sqlite3.Error as er:
    print("Error", er)
# finally: пока не трогать) или делать доступ в бд везде
#     # cursor.close()
#     # db.close()

# Загрузка шаблона Jinja2
env = Environment(
    loader=FileSystemLoader('templates'),
    autoescape=select_autoescape(['html'])
)


# ((((((((((((((((((Обработка HTTP-запросов))))))))))))))
class RequestHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        if self.path == '/':
            self.path = '/users'

        if self.path == '/users':
            # Получение данных из базы данных
            cursor.execute("SELECT * FROM users")
            users = cursor.fetchall()

            # Рендер шаблона с использованием Jinja2
            template = env.get_template('users.html')
            rendered_template = template.render(users=users)

            # Отправка ответа
            self.send_response(200)
            self.send_header('Content-type', 'text/html')
            self.end_headers()
            self.wfile.write(rendered_template.encode())

        elif self.path.endswith('.css'):
            # Отправка CSS-файла
            css_path = os.path.join(os.getcwd(), 'static', 'style.css')
            with open(css_path, 'rb') as css_file:
                self.send_response(200)
                self.send_header('Content-type', 'text/css')
                self.end_headers()
                self.wfile.write(css_file.read())

        elif self.path.endswith('.js'):
            # Отправка JS-файла
            js_path = os.path.join(os.getcwd(), 'static', 'script.js')
            with open(js_path, 'rb') as js_file:
                self.send_response(200)
                self.send_header('Content-type', 'application/javascript')
                self.end_headers()
                self.wfile.write(js_file.read())

        else:
            self.send_response(404)
            self.send_header('Content-type', 'text/html')
            self.end_headers()
            self.wfile.write(b'404 Not Found')

        # ++++++++++++++++++++++++++++++POST________________________________________

    def do_POST(self):
        if self.path == '/users':
            # Чтение данных из тела запроса
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length).decode('utf-8')
            post_data = post_data.split('&')

            regions = ('Краснодарский край', 'Ростовская область', 'Ставропольский край')
            city = ("Краснодар", "Кропоткин", "Славянск", "Ростов", "Шахты", "Батайск",
                    "Ставрополь", "Пятигорск", "Кисловодск")

            # Разделение данных на поля
            second_name = unquote_plus(post_data[0].split('=')[1])
            first_name = unquote_plus(post_data[1].split('=')[1])
            patronymic = unquote_plus(post_data[2].split('=')[1])
            region_id = unquote_plus(post_data[3].split('=')[1])
            print(region_id)
            if region_id == regions[0]:
                region_id = 0
            elif region_id == regions[1]:
                region_id = 1
            elif region_id == regions[2]:
                region_id = 2

            city_id = unquote_plus(post_data[4].split('=')[1])
            city_id = city.index(city_id)

            phone = unquote_plus(post_data[5].split('=')[1])
            email = unquote_plus(post_data[6].split('=')[1])

            # Вставка данных в базу данных
            cursor.execute(
                "INSERT INTO users (second_name, first_name, patronymic, region_id, city_id, phone, email) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (second_name, first_name, patronymic, region_id, city_id, phone, email))
            db.commit()

            # Перенаправление на страницу с пользователями
            self.send_response(303)
            self.send_header('Location', '/users')
            self.end_headers()

        # ((((((((((((((((start import xlsx)))))))))))))))))))))))))

        elif self.path == '/import_xlsx':
            content_length = int(self.headers['Content-Length'])
            body = self.rfile.read(content_length)

            # Сохраняем полученный файл на сервере
            with open('imported_file.xlsx', 'wb') as file:
                file.write(body)

            cursor.execute('''SELECT region_name, id from regions''')
            region = cursor.fetchall()
            print(region)
            cursor.execute('''SELECT city_name, id FROM cities''')
            city = cursor.fetchall()
            print(city)

            # Открываем файл с помощью openpyxl
            workbook = openpyxl.load_workbook('imported_file.xlsx')
            sheet = workbook.active

            # Перебираем строки в файле и добавляем данные в базу данных
            for row in sheet.iter_rows(min_row=2, values_only=True):
                second_name, first_name, patronymic, region_id, city_id, phone, email = row
                print(f'region {region_id}')
                print(f'city {city_id}')

                if region_id in [r[0] for r in region]:
                    region_id = region[[r[0] for r in region].index(region_id)][1] + 1
                    print(region_id)

                if city_id in [c[0] for c in city]:
                    city_id = city[[c[0] for c in city].index(city_id)][1] + 1
                    print(city_id)

                cursor.execute(
                    "INSERT INTO users (second_name, first_name, patronymic, region_id, city_id, phone, email) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (second_name, first_name, patronymic, region_id, city_id, phone, email))

            # Сохраняем изменения
            db.commit()
            # Удаляем файл с сервера
            os.remove('imported_file.xlsx')

            # Редирект на страницу "http://127.0.0.1:8080/users"
            self.send_response(302)
            self.send_header('Location', '/users')
            self.end_headers()

        # ((((((((((((((((end import xlsx)))))))))))))))))))))))))

        # ((((((((((((((((start export xlsx)))))))))))))))))))))))))

        elif self.path == '/export_xlsx':
            # Получение данных из базы данных
            cursor.execute("SELECT * FROM users")
            users = cursor.fetchall()

            # Создание нового документа xlsx
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Запись названия колонок таблицы
            column_names = [i[0] for i in cursor.description]
            for col, column_name in enumerate(column_names, start=1):
                sheet.cell(row=1, column=col, value=column_name)

            # Запись данных в документ xlsx из users
            for row, user in enumerate(users, start=2):
                sheet.cell(row=row, column=1, value=user[0])
                sheet.cell(row=row, column=2, value=user[1])
                sheet.cell(row=row, column=3, value=user[2])
                sheet.cell(row=row, column=4, value=user[3])
                # sheet.cell(row=row, column=5, value=user[4])
                # sheet.cell(row=row, column=6, value=user[5])
                sheet.cell(row=row, column=7, value=user[6])
                sheet.cell(row=row, column=8, value=user[7])

            cursor.execute('''SELECT r.region_name
                              FROM users AS u
                              INNER JOIN regions AS r ON u.region_id = r.id;''')
            region_citi = cursor.fetchall()
            # Запись данных в документ xlsx из regions
            for row, reg_ci in enumerate(region_citi, start=2):
                sheet.cell(row=row, column=5, value=reg_ci[0])

            cursor.execute('''SELECT c.city_name
                              FROM cities AS c 
                              INNER JOIN users AS u ON u.city_id = c.id''')
            cities = cursor.fetchall()
            # Запись данных в документ xlsx из cities
            for row, city in enumerate(cities, start=2):
                sheet.cell(row=row, column=6, value=city[0])

            # Сохранение документа xlsx
            workbook.save('users.xlsx')

            # Отправка файла пользователю
            with open('users.xlsx', 'rb') as xlsx_file:
                self.send_response(200)
                self.send_header('Content-type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', 'attachment; filename="users.xlsx"')
                self.end_headers()
                self.wfile.write(xlsx_file.read())

            # Удаление файла xlsx
            os.remove('users.xlsx')

        # ((((((((((((((((end export xlsx)))))))))))))))))))))))))

        # ((((((((((((((((start import PDF)))))))))))))))))))))))))

        elif self.path == '/import_pdf':
            content_length = int(self.headers['Content-Length'])
            body = self.rfile.read(content_length)

            # Сохраняем полученный файл на сервере
            with open('imported_file.pdf', 'wb') as file:
                file.write(body)

            file_name = 'imported_file.pdf'
            pdf = fitz.open(file_name)

            for page_num in range(1):
                page = pdf[page_num]
                text = page.get_text()

                # Получаем нужную строку текста на странице и тд
                first_1 = text.split('\n')[0]
                first_1 = first_1.split(' ')
                second_name, first_name, patronymic = first_1[0], first_1[1], first_1[2]
                first_2 = text.split('\n')[2]
                first_2 = first_2[0:17]
                first_2 = ''.join(filter(str.isdigit, first_2))
                phone = f"+7 {first_2[1:4]} {first_2[4:7]} {first_2[7:9]} {first_2[9:11]}"
                first_3 = text.split('\n')[3]
                email = first_3
                first_4 = text.split('\n')[4]
                first_4 = first_4.split(' ')
                city = first_4[1]
                if city == 'Краснодар':
                    region_id, city_id, = 0, 0
                cursor.execute(
                    "INSERT INTO users (second_name, first_name, patronymic, region_id, city_id, phone, email) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (second_name, first_name, patronymic, region_id, city_id, phone, email))

            # Сохраняем изменения и закрываем соединение с базой данных
            db.commit()
            # Закрываем файл
            pdf.close()

            # Редирект на страницу "http://127.0.0.1:8080/users"
            self.send_response(302)
            self.send_header('Location', '/users')
            self.end_headers()

            # удаление файла
            os.remove('imported_file.pdf')

        # ((((((((((((((((end import PDF)))))))))))))))))))))))))

        # ((((((((((((((((start export PDF)))))))))))))))))))))))))

        elif self.path == '/export_pdf':
            pdfmetrics.registerFont(TTFont('DejaVuSans', 'static/fonts/DejaVuSans.ttf', 'utf-8'))
            # Получение данных из базы данных
            cursor.execute('''SELECT u.id, u.second_name, u.first_name, u.patronymic, r.region_name, c.city_name, u.phone, u.email
                              FROM users AS u
                              JOIN regions AS r ON u.region_id = r.id
                              JOIN cities AS c ON u.city_id = c.id;''')
            users = cursor.fetchall()

            # Создание PDF файла с резюме
            pdf_file = 'users.pdf'
            c = canvas.Canvas(pdf_file, pagesize=letter)

            # Запись данных в PDF-файл
            for user in users:
                # Настройка шрифта и размера текста
                c.setFont("DejaVuSans", 12)
                # Сами данные
                c.drawString(100, 700, f"ID: {user[0]}")
                c.drawString(100, 680, f"Фамилия: {user[1]}")
                c.drawString(100, 660, f"Имя: {user[2]}")
                c.drawString(100, 640, f"Отчество: {user[3]}")
                c.drawString(100, 620, f"Регион: {user[4]}")
                c.drawString(100, 600, f"Город: {user[5]}")
                c.drawString(100, 580, f"Телефон: {user[6]}")
                c.drawString(100, 560, f"Email: {user[7]}")
                c.showPage()

            c.save()

            # Отправка PDF-файла
            with open(pdf_file, 'rb') as pdf_file:
                self.send_response(200)
                self.send_header('Content-type', 'application/pdf')
                self.send_header('Content-Disposition', 'attachment; filename="users.pdf"')
                self.end_headers()
                self.wfile.write(pdf_file.read())

            # Удаление файла
            os.remove('users.pdf')

        # ((((((((((((((((end export PDF)))))))))))))))))))))))))

        else:
            self.send_response(404)
            self.send_header('Content-type', 'text/html')
            self.end_headers()
            self.wfile.write(b'404 Not Found')


# Запуск сервера
def run_server():
    server_address = ('', 8080)
    httpd = HTTPServer(server_address, RequestHandler)
    print('Start server port 8080')
    httpd.serve_forever()


if __name__ == '__main__':
    run_server()
